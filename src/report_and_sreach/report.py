# src/report_and_sreach/report.py
from __future__ import annotations

from src.utils.time_zone import VN_TZ
from typing import Any, Dict, List, Optional, Iterable, Union
from pathlib import Path
from datetime import datetime, timedelta
import csv
import json
import logging
from collections import defaultdict

from src.utils.validators import parse_iso_datetime

# Optional dependency cho Excel export
try:
    from openpyxl import Workbook
    _HAS_OPENPYXL = True
except ImportError:
    Workbook = None
    _HAS_OPENPYXL = False

# Logger riêng cho module
logger = logging.getLogger(__name__)

# Hằng số
DEFAULT_ENCODING = "utf-8"
DEFAULT_LOOKBACK_DAYS = 30
PREDICT_SOON_DAYS = 7  # số ngày coi như "sắp hết hàng"

# Types
AlertProduct = Dict[str, Any]
Alerts = Dict[str, Any]


# ==============================
# 🔧 Helper Functions
# ==============================

def safe_get(obj: Any, attr: str, default: Any = None) -> Any:
    """Truy cập attr từ object hoặc dict an toàn."""
    if obj is None:
        return default
    return obj.get(attr, default) if isinstance(obj, dict) else getattr(obj, attr, default)


def compute_sale_rates(
    transactions: Iterable[Dict[str, Any]],
    product_ids: Optional[Union[str, Iterable[str]]] = None,
    lookback_days: int = DEFAULT_LOOKBACK_DAYS,
) -> Union[float, Dict[str, float]]:
    """
    Tính tốc độ bán trung bình (số lượng/ngày).

    - product_ids = None → trả về dict cho tất cả
    - product_ids = str  → trả về float cho sản phẩm đó
    - product_ids = list → trả về dict {pid: rate}
    """
    now = datetime.now(VN_TZ)
    start = now - timedelta(days=lookback_days)
    totals: Dict[str, float] = defaultdict(float)

    # Chuẩn hóa product_ids
    single_mode = isinstance(product_ids, str)
    pid_set = {product_ids} if single_mode else (set(product_ids) if product_ids else None)

    for t in transactions:
        if str(safe_get(t, "type", "OUT")).upper() not in ("OUT", "EXPORT"):
            continue

        pid = str(safe_get(t, "product_id") or "").strip()
        if not pid or (pid_set and pid not in pid_set):
            continue

        dt = parse_iso_datetime(safe_get(t, "date"), default_now=False)
        if not dt:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=VN_TZ)

        if dt >= start:
            totals[pid] += float(safe_get(t, "quantity", 0) or 0)

    rates = {pid: round(total / max(1.0, lookback_days), 2) for pid, total in totals.items()}

    return rates.get(product_ids, 0.0) if single_mode else rates


# ==============================
# 📊 Generate Alerts
# ==============================

def generate_low_stock_alerts(
    product_mgr,
    transaction_mgr: Optional[Any] = None,
    *,
    include_out_of_stock: bool = True,
    include_low_stock: bool = True,
    reorder_buffer: int = 0,
    lookback_days: int = DEFAULT_LOOKBACK_DAYS,
) -> Alerts:
    """Sinh cảnh báo tồn kho + dự báo."""
    now_iso = datetime.now(VN_TZ).isoformat()
    out_of_stock, low_stock = [], []
    total_needed = 0
    category_summary: Dict[str, Dict[str, int]] = {}

    # Lấy transactions nếu có
    transactions_all: List[Dict[str, Any]] = []
    if transaction_mgr:
        try:
            if hasattr(transaction_mgr, "list_transactions"):
                transactions_all = list(transaction_mgr.list_transactions())
            elif hasattr(transaction_mgr, "transactions"):
                transactions_all = list(transaction_mgr.transactions)
        except Exception as e:
            logger.warning("Không thể lấy transactions: %s", e)

    sale_rates = compute_sale_rates(transactions_all, product_ids=None, lookback_days=lookback_days) if transactions_all else {}
    if not isinstance(sale_rates, dict):
        sale_rates = {}

    # Duyệt sản phẩm
    for p in product_mgr.list_products():
        pid = str(safe_get(p, "product_id"))
        qty = int(safe_get(p, "stock_quantity", 0) or 0)
        min_thr = int(safe_get(p, "min_threshold", 0) or 0)

        item: AlertProduct = {
            "product_id": pid,
            "name": safe_get(p, "name", ""),
            "category": safe_get(p, "category", "UNCATEGORIZED"),
            "stock_quantity": qty,
            "min_threshold": min_thr,
            "unit": safe_get(p, "unit", ""),
            "needed": max(0, (min_thr - qty) + max(0, reorder_buffer)) if qty < min_thr else 0,
        }

        # Thêm dự báo
        rate = sale_rates.get(pid)
        if rate:
            item["daily_sale_rate"] = rate
            days_left = qty / rate if rate > 0 else None
            if days_left:
                item["days_until_stockout"] = round(days_left, 2)
                item["predicted_out_soon"] = days_left <= PREDICT_SOON_DAYS

        # Phân loại
        cs = category_summary.setdefault(item["category"], {"out_of_stock": 0, "low_stock": 0})
        if qty == 0 and include_out_of_stock:
            out_of_stock.append(item)
            total_needed += item["needed"]
            cs["out_of_stock"] += 1
        elif 0 < qty <= min_thr and include_low_stock:
            low_stock.append(item)
            total_needed += item["needed"]
            cs["low_stock"] += 1

    return {
        "generated_at": now_iso,
        "out_of_stock": out_of_stock,
        "low_stock": low_stock,
        "total_needed": total_needed,
        "by_category": category_summary,
    }


# ==============================
# 📝 Format Reports
# ==============================

def format_alerts_text(alerts: Alerts) -> str:
    """Xuất cảnh báo tồn kho dạng text."""
    lines: List[str] = [
        "========== CẢNH BÁO TỒN KHO ==========",
        f"Thời điểm tạo báo cáo: {alerts.get('generated_at')}",
        "",
    ]

    def _section(title: str, items: List[AlertProduct], icon: str) -> List[str]:
        if not items:
            return [f"{icon} {title}: Không có.", ""]
        section = [f"{icon} {title} ({len(items)} sản phẩm):"]
        for p in items:
            base = f"- {p['product_id']}: {p['name']} ({p['stock_quantity']}/{p['min_threshold']} {p.get('unit','')}) - Cần nhập: {p['needed']}"
            extras = []
            if "daily_sale_rate" in p:
                extras.append(f"bán/ngày={p['daily_sale_rate']}")
            if "days_until_stockout" in p:
                extras.append(f"dự báo hết={p['days_until_stockout']} ngày")
            if extras:
                base += " | " + "; ".join(extras)
            section.append(base)
        return section + [""]

    lines.extend(_section("HẾT HÀNG", alerts.get("out_of_stock", []), "🚨"))
    lines.extend(_section("SẮP HẾT HÀNG", alerts.get("low_stock", []), "⚠️"))

    lines.append(f"Tổng cần nhập (tất cả sản phẩm): {alerts.get('total_needed', 0)}")
    lines.append("")

    if alerts.get("by_category"):
        lines.append("📂 Thống kê theo danh mục:")
        for cat, stats in alerts["by_category"].items():
            lines.append(f"- {cat}: Hết hàng={stats.get('out_of_stock',0)}; Sắp hết={stats.get('low_stock',0)}")

    return "\n".join(lines)


# ==============================
# 💾 Export Functions
# ==============================

def write_low_stock_alerts(
    alerts: Alerts,
    *,
    out_txt_path: Optional[str] = "data/low_stock_alert.txt",
    out_csv_path: Optional[str] = None,
    encoding: str = DEFAULT_ENCODING,
) -> None:
    """Ghi báo cáo ra TXT và CSV."""
    if out_txt_path:
        path = Path(out_txt_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(format_alerts_text(alerts), encoding=encoding)

    if out_csv_path:
        path = Path(out_csv_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "product_id", "name", "category", "stock_quantity", "min_threshold",
            "unit", "needed", "type", "daily_sale_rate", "days_until_stockout", "predicted_out_soon"
        ]
        with path.open("w", encoding=encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for label, items in (("OUT_OF_STOCK", alerts.get("out_of_stock", [])),
                                 ("LOW_STOCK", alerts.get("low_stock", []))):
                for item in items:
                    row = {k: item.get(k, "") for k in fieldnames}
                    row["type"] = label
                    writer.writerow(row)


def alerts_to_json(alerts: Alerts, *, ensure_ascii: bool = False) -> str:
    """Trả về chuỗi JSON (phục vụ API/web)."""
    return json.dumps(alerts, ensure_ascii=ensure_ascii, indent=2, default=str)


def export_alerts_xlsx(alerts: Alerts, out_xlsx_path: str = "data/low_stock_alert.xlsx") -> None:
    """Xuất alerts sang Excel (3 sheet)."""
    if not _HAS_OPENPYXL or Workbook is None:
        raise RuntimeError("openpyxl chưa được cài.")

    wb = Workbook()
    headers = ["product_id", "name", "category", "stock_quantity", "min_threshold",
               "unit", "needed", "daily_sale_rate", "days_until_stockout", "predicted_out_soon"]

    # Out of stock
    ws_oos = wb.active
    if ws_oos is None:
        ws_oos = wb.create_sheet("OutOfStock")
        wb.active = ws_oos
    ws_oos.title = "OutOfStock"
    ws_oos.append(headers)
    for item in alerts.get("out_of_stock", []):
        ws_oos.append([item.get(h, "") for h in headers])

    # Low stock
    ws_low = wb.create_sheet("LowStock")
    ws_low.append(headers)
    for item in alerts.get("low_stock", []):
        ws_low.append([item.get(h, "") for h in headers])

    # By category
    ws_cat = wb.create_sheet("ByCategory")
    ws_cat.append(["category", "out_of_stock", "low_stock"])
    for cat, stats in alerts.get("by_category", {}).items():
        ws_cat.append([cat, stats.get("out_of_stock", 0), stats.get("low_stock", 0)])

    Path(out_xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_xlsx_path))
# ==============================
# 📝 Transaction Log Reports
# ==============================

def format_transaction_log(transactions: List[Any]) -> str:
    """Xuất nhật ký giao dịch dạng text."""
    lines = [
        "========== NHẬT KÝ GIAO DỊCH ==========",
        f"Tổng số giao dịch: {len(transactions)}",
        "",
    ]
    for t in transactions:
        lines.append(
            f"- {t.transaction_id}: {t.product_id} | {t.trans_type} | "
            f"{t.quantity} | {t.date.isoformat()} | {t.note or ''}"
        )
    return "\n".join(lines)


def export_transaction_log(
    transaction_mgr,
    *,
    out_txt_path: Optional[str] = "data/transaction_log.txt",
    out_csv_path: Optional[str] = "data/transaction_log.csv",
    out_json_path: Optional[str] = "data/transaction_log.json",
    out_xlsx_path: Optional[str] = "data/transaction_log.xlsx",
    encoding: str = DEFAULT_ENCODING,
) -> None:
    """Xuất nhật ký giao dịch ra TXT, CSV, JSON, Excel."""

    transactions = transaction_mgr.list_transactions()
    rows = [t.to_dict() for t in transactions]

    # TXT
    if out_txt_path:
        Path(out_txt_path).parent.mkdir(parents=True, exist_ok=True)
        Path(out_txt_path).write_text(format_transaction_log(transactions), encoding=encoding)

    # CSV
    if out_csv_path:
        Path(out_csv_path).parent.mkdir(parents=True, exist_ok=True)
        fieldnames = ["transaction_id", "product_id", "trans_type", "quantity", "date", "note"]
        with Path(out_csv_path).open("w", encoding=encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

    # JSON
    if out_json_path:
        Path(out_json_path).parent.mkdir(parents=True, exist_ok=True)
        Path(out_json_path).write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str), encoding=encoding)

    # Excel
    if out_xlsx_path:
        if not _HAS_OPENPYXL or Workbook is None:   
            logger.warning("openpyxl chưa được cài, bỏ qua export Excel.")
        else:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("Transactions")
                wb.active = ws
            ws.title = "Transactions"
            ws.append(["transaction_id", "product_id", "trans_type", "quantity", "date", "note"])
            for row in rows:
                ws.append([row.get("transaction_id"), row.get("product_id"),
                           row.get("trans_type"), row.get("quantity"),
                           row.get("date"), row.get("note")])
            Path(out_xlsx_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(out_xlsx_path))

# ==============================
# 🚀 Main Entry
# ==============================

def run_and_persist(
    product_mgr,
    transaction_mgr: Optional[Any] = None,
    *,
    reorder_buffer: int = 0,
    lookback_days: int = DEFAULT_LOOKBACK_DAYS,
    out_txt_path: str = "data/low_stock_alert.txt",
    out_csv_path: Optional[str] = "data/low_stock_alert.csv",
    out_xlsx_path: Optional[str] = "data/low_stock_alert.xlsx",
) -> Alerts:
    """Hàm chính: sinh cảnh báo và ghi ra file .txt, .csv, .xlsx."""
    alerts = generate_low_stock_alerts(
        product_mgr,
        transaction_mgr=transaction_mgr,
        reorder_buffer=reorder_buffer,
        lookback_days=lookback_days,
    )
    write_low_stock_alerts(alerts, out_txt_path=out_txt_path, out_csv_path=out_csv_path)
    if out_xlsx_path:
        try:
            export_alerts_xlsx(alerts, out_xlsx_path)
        except Exception as e:
            logger.warning("Không thể export Excel: %s", e)
    return alerts


# ==============================
# Financial / Sales Summary
# ==============================


def compute_financial_summary(
    product_mgr,
    transaction_mgr,
    *,
    top_k: int = 5,
    include_zero_sales: bool = False,
    currency: str = "VND",
    month: Optional[int] = None,
    year: Optional[int] = None,
    out_dir: Optional[Union[str, Path]] = None,
) -> Dict[str, Any]:
    """Compute revenue/profit totals, by-category breakdown and top-K lists.

    If month and year are provided, only transactions in that month/year are counted
    and a CSV file named sales_summary_MM_YYYY.csv will be written to out_dir (or ./data).

    Assumptions:
      - Transactions with trans_type 'IMPORT' are purchases (cost incurred).
      - Transactions with trans_type 'EXPORT' are sales (revenue earned).
      - Product.cost_price and sell_price are Decimal-like strings in Product model.

    Returns a dict with keys: total_revenue, total_cost, total_profit, by_category, top_sellers, least_purchased
    """
    from decimal import Decimal

    # map products by id for quick lookup
    products = {p.product_id: p for p in product_mgr.list_products()}

    total_revenue = Decimal("0")
    total_cost = Decimal("0")
    qty_by_product: Dict[str, int] = defaultdict(int)

    # iterate transactions
    transactions = list(transaction_mgr.list_transactions()) 

    # optionally filter by month/year
    def _in_period(dt) -> bool:
        if dt is None:
            return False
        # dt may be datetime or string
        if not isinstance(dt, datetime):
            dt = parse_iso_datetime(dt, default_now=False)
            if not dt:
                return False
        if month is None or year is None:
            return True
        return (dt.month == int(month)) and (dt.year == int(year))

    for t in transactions:
        # get transaction date and skip if not in requested period
        tdate = getattr(t, "date", None)
        if (month is not None and year is not None) and (not _in_period(tdate)):
            continue

        pid = str(getattr(t, "product_id", "") or "").strip()
        ttype = str(getattr(t, "trans_type", "")).upper()
        qty = int(getattr(t, "quantity", 0) or 0)
        if not pid or qty <= 0:
            continue

        prod = products.get(pid)
        if ttype in ("EXPORT", "OUT"):
            # revenue = sell_price * qty; cost = cost_price * qty
            if prod is not None:
                sell = Decimal(str(prod.sell_price))
                cost = Decimal(str(prod.cost_price))
            else:
                # unknown product: treat prices as 0
                sell = Decimal("0")
                cost = Decimal("0")
            total_revenue += sell * qty
            total_cost += cost * qty
            qty_by_product[pid] += qty
        elif ttype == "IMPORT":
            # purchase increases inventory cost but not revenue; treat as cost
            if prod is not None:
                cost = Decimal(str(prod.cost_price))
            else:
                cost = Decimal("0")
            total_cost += cost * qty
            # do not count imports in sold quantities

    total_profit = total_revenue - total_cost

    # by-category aggregation
    by_category: Dict[str, Dict[str, Any]] = {}
    for pid, sold_qty in qty_by_product.items():
        prod = products.get(pid)
        cat = prod.category if prod is not None else "UNCATEGORIZED"
        if cat not in by_category:
            by_category[cat] = {"revenue": Decimal("0"), "cost": Decimal("0"), "profit": Decimal("0"), "quantity": 0}
        sell = Decimal(str(prod.sell_price)) if prod is not None else Decimal("0")
        cost = Decimal(str(prod.cost_price)) if prod is not None else Decimal("0")
        by_category[cat]["revenue"] += sell * sold_qty
        by_category[cat]["cost"] += cost * sold_qty
        by_category[cat]["profit"] += (sell - cost) * sold_qty
        by_category[cat]["quantity"] += sold_qty

    # convert Decimals to strings for JSON-friendly output
    def dec_map(d: Dict[str, Any]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for k, v in d.items():
            out[k] = str(v) if isinstance(v, Decimal) else v
        return out

    by_category_out = {cat: dec_map(stats) for cat, stats in by_category.items()}

    # top K sellers and least purchased
    items = [(pid, qty) for pid, qty in qty_by_product.items()]
    items_sorted_desc = sorted(items, key=lambda x: x[1], reverse=True)
    items_sorted_asc = sorted(items, key=lambda x: x[1])

    def make_item_list(items_list):
        res = []
        for pid, qty in items_list[:top_k]:
            p = products.get(pid)
            if p is not None:
                revenue = Decimal(str(p.sell_price)) * qty
                cost = Decimal(str(p.cost_price)) * qty
                profit = revenue - cost
            else:
                revenue = Decimal("0")
                cost = Decimal("0")
                profit = Decimal("0")
            res.append({
                "product_id": pid,
                "name": p.name if p is not None else "",
                "category": p.category if p is not None else "",
                "quantity_sold": qty,
                "revenue": str(revenue),
                "cost": str(cost),
                "profit": str(profit),
            })
        return res

    top_sellers = make_item_list(items_sorted_desc)
    least_purchased = make_item_list(items_sorted_asc if include_zero_sales else [it for it in items_sorted_asc if it[1] > 0])

    summary = {
        "total_revenue": str(total_revenue),
        "total_cost": str(total_cost),
        "total_profit": str(total_profit),
        "currency": currency,
        "by_category": by_category_out,
        "top_sellers": top_sellers,
        "least_purchased": least_purchased,
        "product_sales": {pid: qty for pid, qty in qty_by_product.items()},
    }

    # Export CSV if both month and year provided
    if month is not None and year is not None:
        try:
            out_dir_path = Path(out_dir) if out_dir else Path("data")
            out_dir_path.mkdir(parents=True, exist_ok=True)
            fname = f"sales_summary_{int(month):02d}_{int(year)}.csv"
            out_path = out_dir_path / fname
            with out_path.open("w", encoding=DEFAULT_ENCODING, newline="") as f:
                writer = csv.writer(f)
                cur = summary.get("currency", "") or ""
                def with_cur(val: Any) -> str:
                    return f"{val} {cur}" if (val is not None and cur) else (str(val) if val is not None else "")
                # Summary header
                writer.writerow(["Key", "Value"])
                writer.writerow(["total_revenue", with_cur(summary["total_revenue"])])
                writer.writerow(["total_cost", with_cur(summary["total_cost"])])
                writer.writerow(["total_profit", with_cur(summary["total_profit"])])
                writer.writerow(["currency", summary["currency"]])
                writer.writerow([])
                # By category
                writer.writerow(["By Category"])
                writer.writerow(["category", "revenue", "cost", "profit", "quantity"])
                for cat, stats in summary["by_category"].items():
                    writer.writerow([
                        cat,
                        with_cur(stats.get("revenue", "0")),
                        with_cur(stats.get("cost", "0")),
                        with_cur(stats.get("profit", "0")),
                        stats.get("quantity", 0),
                    ])
                writer.writerow([])
                # Top sellers
                writer.writerow(["Top Sellers"])
                writer.writerow(["name", "category", "quantity_sold", "revenue", "cost", "profit"])
                for it in summary["top_sellers"]:
                    writer.writerow([
                        it.get("name", ""),
                        it.get("category", ""),
                        it.get("quantity_sold", 0),
                        with_cur(it.get("revenue", "0")),
                        with_cur(it.get("cost", "0")),
                        with_cur(it.get("profit", "0")),
                    ])
                writer.writerow([])
                # Least purchased
                writer.writerow(["Least Purchased"])
                writer.writerow(["name", "category", "quantity_sold", "revenue", "cost", "profit"])
                for it in summary["least_purchased"]:
                    writer.writerow([
                        it.get("name", ""),
                        it.get("category", ""),
                        it.get("quantity_sold", 0),
                        with_cur(it.get("revenue", "0")),
                        with_cur(it.get("cost", "0")),
                        with_cur(it.get("profit", "0")),
                    ])
            logger.info("Wrote sales summary CSV: %s", str(out_path))
        except Exception as e:
            logger.exception("Failed to write sales summary CSV: %s", e)

    return summary


def format_financial_summary_text(summary: Dict[str, Any]) -> str:
    lines = ["=== Financial Summary ==="]
    cur = summary.get("currency") or ""
    def with_currency(val: Any) -> str:
        if val is None:
            return ""
        return f"{val} {cur}" if cur else str(val)

    lines.append(f"Total Revenue: {with_currency(summary.get('total_revenue'))}")
    lines.append(f"Total Cost: {with_currency(summary.get('total_cost'))}")
    lines.append(f"Total Profit: {with_currency(summary.get('total_profit'))}")
    lines.append("")
    lines.append("By Category:")
    for cat, stats in summary.get("by_category", {}).items():
        rev = with_currency(stats.get('revenue'))
        cost = with_currency(stats.get('cost'))
        prof = with_currency(stats.get('profit'))
        lines.append(f"- {cat}: revenue={rev} cost={cost} profit={prof} quantity={stats.get('quantity')}")
    lines.append("")
    lines.append("Top Sellers:")
    for it in summary.get("top_sellers", []):
        rev = with_currency(it.get("revenue"))
        cost = with_currency(it.get("cost"))
        prof = with_currency(it.get("profit"))
        lines.append(f"- {it.get('name','')} ({it.get('category','')}) quantity={it.get('quantity_sold',0)} revenue={rev} cost={cost} profit={prof}")
    lines.append("")
    lines.append("Least Purchased:")
    for it in summary.get("least_purchased", []):
        rev = with_currency(it.get("revenue"))
        cost = with_currency(it.get("cost"))
        prof = with_currency(it.get("profit"))
        lines.append(f"- {it.get('name','')} ({it.get('category','')}) quantity={it.get('quantity_sold',0)} revenue={rev} cost={cost} profit={prof}")
    return "\n".join(lines)